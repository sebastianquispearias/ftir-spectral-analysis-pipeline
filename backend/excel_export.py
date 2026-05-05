from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from backend.anova_analysis import construir_matriz_diseno_box_behnken


def generar_excel(
    df_resultados: pd.DataFrame,
    anova_resultado: dict | None,
    ruta_salida: str | Path,
    anchor_points: list[float] | None = None,
    rangos: dict[str, tuple[float, float]] | None = None,
) -> str:
    """Generate an Excel workbook with all analysis results."""
    ruta_salida = Path(ruta_salida)
    wb = Workbook()

    _hoja_datos_crudos(wb, df_resultados)
    _hoja_resumen(wb, df_resultados)
    if anova_resultado:
        _hoja_anova(wb, anova_resultado)
        _hoja_coeficientes(wb, anova_resultado)
    _hoja_info(wb, anchor_points, rangos)
    _hoja_anchor_config(wb, anchor_points)

    del wb["Sheet"]
    wb.save(str(ruta_salida))
    return str(ruta_salida)


def _estilo_header() -> tuple[Font, Border, Alignment]:
    font = Font(bold=True, size=11)
    border = Border(bottom=Side(style="thin"))
    alignment = Alignment(horizontal="center")
    return font, border, alignment


def _autofit_columns(ws) -> None:
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 30)


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    font, border, alignment = _estilo_header()
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font
        cell.border = border
        cell.alignment = alignment


def _hoja_datos_crudos(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Datos Crudos")
    headers = ["Archivo", "Experimento", "Réplica", "Altura Carboxilato",
               "Área Carboxilato", "Normalizada", "Altura Referencia", "Posición Pico (cm⁻¹)"]
    _write_header(ws, headers)

    cols = ["archivo", "experimento", "replica", "altura_carb", "area_carb",
            "normalizada", "altura_ref", "x_pico_carb"]

    for i, (_, row) in enumerate(df.iterrows(), 2):
        for j, col in enumerate(cols, 1):
            val = row[col]
            cell = ws.cell(row=i, column=j, value=val)
            if isinstance(val, float):
                cell.number_format = "0.0000"

    _autofit_columns(ws)


def _hoja_resumen(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Resumen por Experimento")

    metricas = ["altura_carb", "area_carb", "normalizada"]
    headers = ["Exp", "Temp (°C)", "Tiempo (min)", "NaClO (mL)"]
    for m in metricas:
        headers.extend([f"{m} (media)", f"{m} (std)"])
    _write_header(ws, headers)

    diseno = construir_matriz_diseno_box_behnken()
    resumen = df.groupby("experimento").agg(
        {m: ["mean", "std"] for m in metricas}
    )

    for i, (_, design_row) in enumerate(diseno.iterrows(), 2):
        exp = int(design_row["exp"])
        ws.cell(row=i, column=1, value=exp)
        ws.cell(row=i, column=2, value=design_row["temperatura"])
        ws.cell(row=i, column=3, value=design_row["tiempo"])
        ws.cell(row=i, column=4, value=design_row["naclo"])

        col_offset = 5
        for m in metricas:
            if exp in resumen.index:
                mean_val = float(resumen.loc[exp, (m, "mean")])
                std_val = float(resumen.loc[exp, (m, "std")])
            else:
                mean_val, std_val = 0.0, 0.0
            ws.cell(row=i, column=col_offset, value=mean_val).number_format = "0.0000"
            ws.cell(row=i, column=col_offset + 1, value=std_val).number_format = "0.0000"
            col_offset += 2

    _autofit_columns(ws)


def _hoja_anova(wb: Workbook, resultado: dict) -> None:
    ws = wb.create_sheet("ANOVA")
    tabla = resultado["tabla_anova"]

    headers = ["Fuente", "Sum Sq", "df", "F", "p-value"]
    _write_header(ws, headers)

    for i, fuente in enumerate(tabla["fuente"], 2):
        idx = i - 2
        ws.cell(row=i, column=1, value=fuente)
        ws.cell(row=i, column=2, value=tabla["sum_sq"][idx]).number_format = "0.000000"
        ws.cell(row=i, column=3, value=tabla["df"][idx]).number_format = "0"
        f_val = tabla["F"][idx]
        ws.cell(row=i, column=4, value=f_val if not pd.isna(f_val) else "").number_format = "0.0000"
        p_val = tabla["PR(>F)"][idx]
        cell = ws.cell(row=i, column=5, value=p_val if not pd.isna(p_val) else "")
        if isinstance(p_val, float) and not pd.isna(p_val):
            cell.number_format = "0.000000"
            if p_val < 0.05:
                cell.font = Font(bold=True, color="FF0000")

    row_extra = len(tabla["fuente"]) + 3
    ws.cell(row=row_extra, column=1, value="R²").font = Font(bold=True)
    ws.cell(row=row_extra, column=2, value=resultado["r_squared"]).number_format = "0.0000"
    ws.cell(row=row_extra + 1, column=1, value="R² ajustado").font = Font(bold=True)
    ws.cell(row=row_extra + 1, column=2, value=resultado["r_squared_adj"]).number_format = "0.0000"

    _autofit_columns(ws)


def _hoja_coeficientes(wb: Workbook, resultado: dict) -> None:
    ws = wb.create_sheet("Coeficientes")
    headers = ["Término", "Coeficiente", "p-value", "Significativo"]
    _write_header(ws, headers)

    for i, (nombre, valor) in enumerate(resultado["coeficientes"].items(), 2):
        ws.cell(row=i, column=1, value=nombre)
        ws.cell(row=i, column=2, value=valor).number_format = "0.000000"
        p = resultado["p_values"].get(nombre, 1.0)
        ws.cell(row=i, column=3, value=p).number_format = "0.000000"
        sig = "Sí" if p < 0.05 else "No"
        cell = ws.cell(row=i, column=4, value=sig)
        if sig == "Sí":
            cell.font = Font(bold=True, color="FF0000")

    _autofit_columns(ws)


def _hoja_info(
    wb: Workbook,
    anchor_points: list[float] | None,
    rangos: dict[str, tuple[float, float]] | None,
) -> None:
    ws = wb.create_sheet("Info")
    info = [
        ("Fecha de generación", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Anchor points", str(anchor_points) if anchor_points else "N/A"),
        ("Rango carboxilato", str(rangos.get("carboxilato")) if rangos else "N/A"),
        ("Rango referencia", str(rangos.get("referencia")) if rangos else "N/A"),
        ("Software", "FTIR Spectral Analysis Pipeline"),
        ("Versión", "1.0.0"),
    ]

    _write_header(ws, ["Parámetro", "Valor"])
    for i, (param, val) in enumerate(info, 2):
        ws.cell(row=i, column=1, value=param).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)

    _autofit_columns(ws)


def _hoja_anchor_config(wb: Workbook, config: dict | None) -> None:
    ws = wb.create_sheet("Anchor Points Config")

    if not config or not isinstance(config, dict):
        ws.cell(row=1, column=1, value="No configuration data available")
        return

    _write_header(ws, ["Parameter", "Value"])
    row = 2

    for key, val in config.items():
        if key == "custom_anchor_points" and val is not None:
            ws.cell(row=row, column=1, value="Mode").font = Font(bold=True)
            ws.cell(row=row, column=2, value="Manual (custom anchor points)")
            row += 1
            ws.cell(row=row, column=1, value="N anchor points").font = Font(bold=True)
            ws.cell(row=row, column=2, value=len(val))
            row += 2
            _write_header(ws, ["Index", "Wavenumber (cm-1)"], row)
            row += 1
            for i, cm in enumerate(sorted(val), 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=cm).number_format = "0.0"
                row += 1
        elif key == "custom_anchor_points" and val is None:
            ws.cell(row=row, column=1, value="Mode").font = Font(bold=True)
            ws.cell(row=row, column=2, value="Automatic (valley detection)")
            row += 1
        else:
            ws.cell(row=row, column=1, value=str(key)).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(val))
            row += 1

    _autofit_columns(ws)
